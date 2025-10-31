# -*- coding: utf-8 -*-
import csv
import json
import logging
import multiprocessing
import os
import random
import sqlite3
import time

import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font
from typing import Dict, List, Any, Optional, Tuple, OrderedDict

from tqdm import tqdm
from transformers import PreTrainedTokenizerBase

from .const import Q_STATIC, Q_DYNAMIC
from .models import SQLTable


def setup_logging(name, log_level=logging.INFO):
    name = Path(name).name
    logger = logging.getLogger('model_performance')

    if log_level == 'debug':
        logger.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s: %(lineno)d - %(message)s')
    else:
        logger.setLevel(logging.INFO)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s: %(message)s')

    fh = logging.FileHandler(f'{name}.log')
    console_handler = logging.StreamHandler()

    fh.setFormatter(formatter)
    logger.addHandler(fh)

    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    return logger


logger = setup_logging('model_performance')


class IOUtils:
    """日志、CSV、Excel 相关工具"""

    @staticmethod
    def get_data_df(sql: str, conn, mode, to_round: int = 2) -> pd.DataFrame:
        try:
            df = pd.read_sql_query(sql, conn).round(to_round)
        except Exception as e:
            raise e
        finally:
            conn.close()

        # 排序
        if mode == 'static':
            # 提取冒号前的数字并转换为整数
            df['label_num'] = df['question_label'].str.split(':').str[0].astype(int)

            # 根据 label_num 和 batch 进行排序
            df_sorted = df.sort_values(by=['label_num', 'batch'])

            # 删除临时列
            df_sorted = df_sorted.drop(columns=['label_num'])
        else:
            df_sorted = df.sort_values(by=['batch', 'num_prompts'])
        return df_sorted

    def export_to_color_excel(self,
                              df: pd.DataFrame, case_id: str, model_id: str, output_dir: Path,
                              label_color: str = 'question_label',
                              color_col_values: tuple = ("64:128", "128:2048", "1024:128", "2048:2048"),
                              label_max: str = 'temp_label_max_batch',
                              mode: str = 'static'
                              ):
        if not output_dir.exists():
            output_dir.mkdir(parents=True, exist_ok=True)

        # 设置填充样式
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        color_fill_0 = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
        color_fill_1 = PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid')
        color_fill_2 = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid')
        color_fill_3 = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')
        color_fill_c = [color_fill_0, color_fill_1, color_fill_2, color_fill_3]

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        font_bold = Font(bold=True)

        # 创建一个新的工作簿和工作表
        new_workbook = Workbook()
        new_sheet = new_workbook.active

        for r in dataframe_to_rows(df, index=False, header=True):
            new_sheet.append(r)

        index_color_col = None
        index_max_col = None

        _color_index = 0
        for row in new_sheet.iter_rows():
            for cell in row:
                if cell.value == label_color:
                    index_color_col = cell.column - 1
                    for c in row:
                        c.fill = yellow_fill
                        c.border = thin_border

                if cell.value == label_max:
                    index_max_col = cell.column - 1

            if index_color_col is not None and row[index_color_col].value in color_col_values:
                _i = color_col_values.index(row[index_color_col].value)
                for c in row:
                    c.fill = color_fill_c[_i % len(color_fill_c)]
                    c.border = thin_border

            if index_max_col is not None and row[index_max_col].value == 1:
                for c in row:
                    c.font = font_bold

        self.delete_cols_remain_specify_cols_name(new_sheet)
        self.auto_adjust_column_width(new_sheet)
        _save_name = '_'.join([mode, Path(model_id).name, time.strftime("%Y%m%d%H%M%S"), str(case_id)])
        new_workbook.save(output_dir / f'{_save_name}.xlsx')

    @staticmethod
    def delete_cols_remain_specify_cols_name(sheet):
        del_col_inds = []

        for row in sheet.iter_rows():
            for cell in row:
                if "temp_" in cell.value:
                    index_question_col = cell.column
                    del_col_inds.append(index_question_col)
            # 拿到第一行表头
            break
        for _ind in sorted(del_col_inds, reverse=True):
            sheet.delete_cols(idx=_ind)

        return sheet

    @staticmethod
    def auto_adjust_column_width(worksheet):
        """
        函数用于自动调整列宽
        :param sheet:
        :return:
        """
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = get_column_letter(column_cells[0].column)
            adjusted_width = length * 1.1
            worksheet.column_dimensions[column_letter].width = adjusted_width

        return worksheet

    @staticmethod
    def find_max_concurrent_under_slo(
            df,
            conditions=('TTFT_P99', 'mean_decode_throughput'),
            thresholds=(5000, 10),
            compare_methods=('<', '>'),
            target='batch',
            tmp_col='temp_label_max_batch',
            group_by='question_label'
    ):
        # 参数校验
        if not (len(conditions) == len(thresholds) == len(compare_methods)):
            raise ValueError("The lengths of conditions, thresholds, and compare_methods must be equal.")

        # 比较运算符映射
        compare_ops = {
            '<': np.less,
            '>': np.greater,
            '<=': np.less_equal,
            '>=': np.greater_equal,
            '==': np.equal,
            '!=': np.not_equal
        }

        # 类型转换（原地修改避免内存拷贝）
        for col in conditions:
            df[col] = pd.to_numeric(df[col], errors='coerce')

        # 向量化条件计算
        conditions_list = [
            compare_ops[method](df[col], threshold)
            for col, method, threshold in zip(conditions, compare_methods, thresholds)
        ]

        # 合并条件（使用logical_and.reduce提升性能）
        filter_mask = np.logical_and.reduce(conditions_list)

        # 分组求极值
        filtered_df = df[filter_mask]
        max_indices = filtered_df.groupby(group_by)[target].idxmax()

        # 使用批量标记（比逐行操作快10倍+）
        df[tmp_col] = df.index.isin(max_indices).astype(int)

        return df


class DatasetLoader:
    def __init__(self, dataset_path, dataset_name, dataset_field=None):
        self.dataset_path = dataset_path
        self.dataset_name = dataset_name
        self.field = dataset_field
        self.dataset = self.load_data()

    def load_data(self):
        logger.info('loading dataset...')
        if self.dataset_path.endswith('.json'):
            with open(self.dataset_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return data
        else:
            raise ValueError("Unsupported file format")

    def _filtered_worker(self, args):
        in_len, out_len, tokenizer, dataset = args
        filtered_data = []

        for i in tqdm(range(len(dataset[in_len])), desc=f"filtering {in_len} {out_len}", leave=False):
            truncated_input = tokenizer(dataset[in_len][i], truncation=True, max_length=out_len)
            truncated_string = tokenizer.decode(truncated_input.input_ids, skip_special_tokens=True)
            filtered_data.append(truncated_string)
        return {in_len: filtered_data}

    def sample_filtered_requests(self, input_len, output_len, tokenizer):
        filtered_dataset = {}
        pool = multiprocessing.Pool()
        tasks = [(in_len, out_len, tokenizer, self.dataset) for in_len, out_len in zip(input_len, output_len)]
        results = pool.map(self._filtered_worker, tasks)
        for res in results:
            filtered_dataset.update(res)
        pool.close()
        pool.join()
        return filtered_dataset

    def sample_requests(self, num_requests, fixed_output_len, shuffle,
                        tokenizer=None, prompt_max_len=None,
                        prompt_len_scale=0, enable_same_prompt=False):
        if self.dataset_name == 'sharegpt':
            logger.info('sampling sharegpt dataset...')
            return self.sample_sharegpt_requests(num_requests, tokenizer, fixed_output_len, shuffle,
                                                 prompt_max_len, prompt_len_scale, enable_same_prompt)
        elif self.dataset_name == 'normal':
            logger.info('sampling normal dataset...')
            return self.sample_normal(num_requests, tokenizer, fixed_output_len, shuffle)
        elif self.dataset_name == 'custom':
            logger.info('sampling custom dataset...')
            if self.field is None:
                return self.sample_custom(num_requests, tokenizer, fixed_output_len, shuffle)
            else:
                return self.sample_custom(num_requests, tokenizer, fixed_output_len, shuffle, self.field)
        else:
            raise ValueError("Unsupported dataset")

    def sample_normal(self, num_requests, tokenizer, fixed_output_len, shuffle):
        if shuffle:
            random.shuffle(self.dataset)

        req_dataset = []
        for prompt in self.dataset:
            req_prompt = prompt['instruction']
            out_prompt = prompt['output']
            if len(req_dataset) == num_requests:
                break
            prompt_token_ids = tokenizer(req_prompt).input_ids
            completion_token_ids = tokenizer(out_prompt).input_ids
            prompt_len = len(prompt_token_ids)
            output_len = len(completion_token_ids
                             ) if fixed_output_len is None else fixed_output_len

            req_dataset.append((req_prompt, prompt_len, output_len, None))

        return req_dataset

    def sample_custom(self, num_requests, tokenizer,
                      fixed_output_len: Optional[int] = None, shuffle: bool = False):
        req_dataset = []
        if self.field is None:
            return req_dataset
        if shuffle:
            random.shuffle(self.dataset)

        for prompt in self.dataset:
            req_prompt = "".join(prompt[in_field] for in_field in self.field['input_field'])
            out_prompt = "".join(prompt[out_field] for out_field in self.field['output_field'])
            if len(req_dataset) == num_requests:
                break
            prompt_token_ids = tokenizer(req_prompt).input_ids
            completion_token_ids = tokenizer(out_prompt).input_ids
            prompt_len = len(prompt_token_ids)
            output_len = len(completion_token_ids
                             ) if fixed_output_len is None else fixed_output_len
            req_dataset.append((req_prompt, prompt_len, output_len, None))

        return req_dataset

    def sample_sharegpt_requests(
            self,
            num_requests: int,
            tokenizer: PreTrainedTokenizerBase,
            fixed_output_len: Optional[int] = None,
            shuffle: bool = False,
            prompt_max_len: Optional[int] = None,
            prompt_len_scale: float = 0,
            enable_same_prompt: bool = False,
    ) -> List[Tuple[str, int, int, None]]:
        # Filter out the conversations with less than 2 turns.
        dataset = [data for data in self.dataset if len(data["conversations"]) >= 2]
        # Only keep the first two turns of each conversation.
        dataset = [(data["conversations"][0]["value"],
                    data["conversations"][1]["value"]) for data in dataset]

        # Shuffle the dataset.
        if shuffle:
            logger.debug("shuffling dataset...")
            random.shuffle(dataset)

        # Filter out sequences that are too long or too short
        filtered_dataset: List[Tuple[str, int, int, None]] = []
        same_prompt_dataset = []
        for i in range(len(dataset)):
            if len(filtered_dataset) == num_requests:
                break

            # Tokenize the prompts and completions.
            prompt = dataset[i][0]
            prompt_token_ids = tokenizer(prompt).input_ids
            completion = dataset[i][1]
            completion_token_ids = tokenizer(completion).input_ids
            prompt_len = len(prompt_token_ids)
            output_len = len(completion_token_ids
                             ) if fixed_output_len is None else fixed_output_len

            if prompt_max_len is None:
                if prompt_len < 4 or (fixed_output_len is None and output_len < 4):
                    # Prune too short sequences.
                    continue
                if prompt_len > 1024 or prompt_len + output_len > 2048:
                    # Prune too long sequences.
                    continue
            else:
                if prompt_len > prompt_max_len + int(prompt_max_len * prompt_len_scale):
                    # 随机截取prompt，长度在prompt_max_len - int(prompt_max_len * prompt_len_scale)
                    # 和 prompt_max_len + int(prompt_max_len * prompt_len_scale)之间
                    prompt_token_ids = prompt_token_ids[
                                       :random.randint(prompt_max_len - int(prompt_max_len * prompt_len_scale),
                                                       prompt_max_len + int(prompt_max_len * prompt_len_scale))]
                    prompt_len = len(prompt_token_ids)
                    prompt = tokenizer.decode(prompt_token_ids, skip_special_tokens=True)
                else:
                    # Prune too short sequences.
                    continue

            if enable_same_prompt:
                # same_prompt_dataset 添加num_requests个相同prompt
                for _ in range(num_requests):
                    same_prompt_dataset.append((prompt, prompt_len, output_len, None))
                return same_prompt_dataset

            filtered_dataset.append((prompt, prompt_len, output_len, None))

        return filtered_dataset


class Result:
    def __init__(self, output_dir):
        self.output_dir = output_dir
        self.mk_result_dir()

    def mk_result_dir(self):
        if not os.path.exists(self.output_dir):
            logger.info(f"mkdir result dir: {self.output_dir}")
            os.makedirs(self.output_dir)

    def gen_create_table_sql(self, table, fields):
        new_field = []
        for field in fields:
            # 对所有field加上``，防止特殊格式导致创建数据表失败
            if "`" not in field:
                split_value = field.split()
                new_field.append("`%s` " % split_value[0] + " ".join(split_value[1:]))
            else:
                new_field.append(field)
        fields = ",".join(str(f) for f in new_field)
        sql = "CREATE TABLE IF NOT EXISTS {0} ({1})".format(table, fields)
        return sql

    def gen_insert_sql(self, table, fields, values):
        sql = """INSERT INTO {0}({1}) VALUES({2})""".format(
            table,
            ",".join("`%s`" % f for f in fields),
            ",".join("'%s'" % v for v in values),
        )
        return sql

    def output_db(self, result_dict, filename, metadata: OrderedDict):
        result_dict = self.combine_save_data(metadata, result_dict)
        mean_prefill = result_dict.get("mean_prefill_throughput", 0)
        mean_decode = result_dict.get("mean_decode_throughput", 0)
        gpu_num = result_dict.get("gpu_num", 1)
        result_dict["prefill_throughput_per_gpu"] = float(mean_prefill) / float(gpu_num)
        result_dict["decode_throughput_per_gpu"] = float(mean_decode) / float(gpu_num)

        dbpath = os.path.join(self.output_dir, filename + '.db')
        conn = sqlite3.connect(dbpath)
        cursor = conn.cursor()
        table_name = "summary_data"

        sql = self.gen_create_table_sql(table_name, SQLTable.sql_summary_title)
        try:
            cursor.execute(sql)

            # 从字典中提取字段和对应的值
            fields = list(result_dict.keys())
            values = list(result_dict.values())

            sql = self.gen_insert_sql(table_name, fields, values)
            cursor.execute(sql)
            conn.commit()
        except Exception as ex:
            logger.error("Execute SQL failed，SQL: {}, error: {}".format(sql, repr(ex)))
        finally:
            cursor.close()
            conn.close()

    def output_csv(self, result_dict, filename, metadata: OrderedDict):
        result_dict = self.combine_save_data(metadata, result_dict)
        mean_prefill = result_dict.get("mean_prefill_throughput", 0)
        mean_decode = result_dict.get("mean_decode_throughput", 0)
        gpu_num = result_dict.get("gpu_num", 1)
        result_dict["prefill_throughput_per_gpu"] = float(mean_prefill) / float(gpu_num)
        result_dict["decode_throughput_per_gpu"] = float(mean_decode) / float(gpu_num)

        filename = Path(filename).name + ".csv"
        result_path = os.path.join(self.output_dir, filename)

        file_exists = os.path.isfile(result_path)
        with open(result_path, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC)
            if not file_exists:
                writer.writerow(result_dict.keys())
            writer.writerow(result_dict.values())

    @staticmethod
    def combine_save_data(metadata: OrderedDict, result: OrderedDict):
        # metadata_dict = OrderedDict()
        # if metadata:
        #     for item in metadata:
        #         key, value = item.split('=')
        #         metadata_dict[key] = value

        metadata.update(result)
        return metadata

    @staticmethod
    def export_to_excel(case_id, model_id, db_path,
                        mode, goodput,
                        tested_input_output_label,
                        result_dir, export_col='', result_dirname='results'):
        logger.info('exporting to excel...')
        # db_path, arch, gpu_type, gpu_count, instance_count, model_id, backend
        if mode == 'static':
            query_sql = Q_STATIC
        elif mode == 'dynamic':
            query_sql = Q_DYNAMIC
        else:
            raise ValueError(f"Invalid mode: {mode}")

        conditions = tuple(goodput.keys())
        thresholds = tuple(goodput.values())

        export_col = ', ' + ', '.join(export_col) if export_col else ''

        _df = IOUtils.find_max_concurrent_under_slo(
            df=IOUtils.get_data_df(
                query_sql.format(case_id=case_id, export_col=export_col),
                conn=sqlite3.connect(db_path),
                mode=mode,
                to_round=2),
            conditions=conditions,
            thresholds=thresholds,
            compare_methods=('<', '<'),
            target='batch',
            tmp_col='temp_label_max_batch',
            group_by='question_label'
        )
        # save to excel
        IOUtils().export_to_color_excel(
            df=_df,
            label_color='question_label',
            color_col_values=tuple(tested_input_output_label),
            label_max='temp_label_max_batch',
            case_id=case_id,
            model_id=model_id,
            output_dir=Path(result_dir) / result_dirname,
            mode=mode
        )


class StopStrategy:
    def __init__(self, stop_strategy, init_acc=False, acc_col="mean_TTFT"):
        self.strategy = stop_strategy
        self.init_acc = init_acc  # 开启精准测试
        self.acc_col = acc_col

        self.tested_batches = {}
        self.accuracy_enabled = self.init_acc
        self.slo_line_arrived = False

    def assert_stop(self):
        if self.strategy == 'slo':
            if self.init_acc:
                stop = self.slo_line_arrived and not self.accuracy_enabled
            else:
                stop = self.slo_line_arrived
        elif self.strategy == 'acc':
            stop = self.accuracy_enabled
        else:
            stop = False
        return stop

    def init_flag(self):
        self.tested_batches = {}
        self.accuracy_enabled = self.init_acc
        self.slo_line_arrived = False

    def at_slo(self, result_dict, stop_slo_dict, input_len, output_len):
        # 计算需要检查的键集合（stop_slo_dict和result_dict的交集）
        keys_to_check = set(stop_slo_dict.keys()) & set(result_dict.keys())
        # 使用any()短路特性快速判断是否存在违规SLO
        slo_violated = any(
            result_dict[key] > stop_slo_dict[key]
            for key in keys_to_check
        )
        # 更新状态标志（使用位或操作保持原有状态）
        self.slo_line_arrived |= slo_violated

        logger.info(f"check SLO flag '{input_len}:{output_len}' {self.slo_line_arrived}")

    @staticmethod
    def OLS(x, y, t_ref=3):
        """
        OLS: 线性回归
        """
        a, b = np.polyfit(x, y, deg=1)
        func = lambda y: (y - b) / a
        return int(np.ceil(func(t_ref)))  # np.ceil().astype(int)实际存的是np.int64(x)

    def _find_keys_pair(self, expected_value):
        less = []
        greater = []
        for batch, value in self.tested_batches.items():
            if value < expected_value:
                less.append(batch)
            elif value > expected_value:
                greater.append(batch)

        left = max(less) if less else None
        right = min(greater) if greater else None

        logger.debug(f"left: {left}, right: {right}, \n{self.tested_batches}")
        if left is not None and right is not None:
            return (left, right), (self.tested_batches[left], self.tested_batches[right])
        else:
            return None, None

    def calculate_accuracy_batch(self, batch, tested_value, expected_value):
        """计算准确性测试批次

        Args:
            batch: 当前批次
            result: 当前批次的结果
            key: 需要检查的键
            expected_value: 预期的准确值

        Returns:
            新计算的批次（如果需要继续测试）或 None（如果测试完成）
        """

        if not self.accuracy_enabled:
            return None

        # 更新测试记录
        self.tested_batches.update({batch: tested_value})
        # 检查当前批次是否满足预期准确性
        if expected_value <= tested_value <= expected_value + expected_value * 0.1:  # 3000 ~ 3300ms
            logger.info(f"Batch {batch} meets the expected accuracy.")
            self.accuracy_enabled = False
            self.tested_batches.clear()
            return None

        x, y = self._find_keys_pair(expected_value)
        # 不存在跨越expected_value的值
        if x is None and y is None:
            return None

        # 存在跨越点的两个值，计算新批次
        calculated_batch = self.OLS(x=x, y=y, t_ref=expected_value)
        logger.debug(f"calculated_batch: {calculated_batch}")
        # 检查新批次是否已存在
        if calculated_batch in self.tested_batches:
            logger.debug(f"Batch {calculated_batch} has been tested. Accuracy test set to False.")
            self.accuracy_enabled = False
            self.tested_batches.clear()
            return None

        logger.info(f"Expect Value: {expected_value}. Attempting to get accuracy batch: {calculated_batch}")
        return calculated_batch
